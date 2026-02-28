"""
tests/test_suite.py
--------------------
Unit and integration tests for climate-risk-transition-monitor.

Run: pytest tests/ -v
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import pytest
import pandas as pd
import numpy as np


# ── NGFSLoader ────────────────────────────────────────────────────────────────

class TestNGFSLoader:

    def test_synthetic_load_returns_ngfsdata(self):
        from src.ngfs_loader import NGFSLoader
        loader = NGFSLoader(use_api=False, use_cache=False)
        data = loader.load()
        assert data is not None
        assert data.source == "synthetic"

    def test_scenarios_present(self):
        from src.ngfs_loader import NGFSLoader
        loader = NGFSLoader(use_api=False, use_cache=False)
        data = loader.load()
        assert "Net Zero 2050" in data.scenarios
        assert "Current Policies" in data.scenarios

    def test_emissions_shape(self):
        from src.ngfs_loader import NGFSLoader
        loader = NGFSLoader(use_api=False, use_cache=False)
        data = loader.load()
        # MultiIndex: (scenario, sector) × years
        assert data.emissions.shape[0] > 0
        assert data.emissions.shape[1] > 0

    def test_carbon_prices_positive(self):
        from src.ngfs_loader import NGFSLoader
        loader = NGFSLoader(use_api=False, use_cache=False)
        data = loader.load()
        assert (data.carbon_prices >= 0).all().all()

    def test_nz2050_higher_price_than_current_policies(self):
        from src.ngfs_loader import NGFSLoader
        loader = NGFSLoader(use_api=False, use_cache=False)
        data = loader.load()
        nz_price_2050 = data.carbon_prices.loc["Net Zero 2050", 2050]
        cp_price_2050 = data.carbon_prices.loc["Current Policies", 2050]
        assert nz_price_2050 > cp_price_2050

    def test_summary_table_columns(self):
        from src.ngfs_loader import NGFSLoader
        loader = NGFSLoader(use_api=False, use_cache=False)
        summary = loader.summary_table()
        assert "Carbon Price 2030 (USD/tCO2)" in summary.columns
        assert "Temperature 2100 (°C)" in summary.columns

    def test_nz2050_emissions_decline(self):
        """NZ2050 total emissions should be lower in 2050 than 2020."""
        from src.ngfs_loader import NGFSLoader
        loader = NGFSLoader(use_api=False, use_cache=False)
        data = loader.load()
        emis_2020 = data.emissions.xs("Net Zero 2050", level="scenario")[2020].sum()
        emis_2050 = data.emissions.xs("Net Zero 2050", level="scenario")[2050].sum()
        assert emis_2050 < emis_2020

    def test_current_policies_emissions_stable_or_growing(self):
        """Current Policies total emissions should not decline rapidly."""
        from src.ngfs_loader import NGFSLoader
        loader = NGFSLoader(use_api=False, use_cache=False)
        data = loader.load()
        emis_2020 = data.emissions.xs("Current Policies", level="scenario")[2020].sum()
        emis_2050 = data.emissions.xs("Current Policies", level="scenario")[2050].sum()
        # Should be within 50% of 2020 level (no dramatic decline)
        assert emis_2050 >= emis_2020 * 0.5


# ── SectorMapper ──────────────────────────────────────────────────────────────

class TestSectorMapper:

    def test_gics_to_ngfs_known(self):
        from src.sector_mapper import map_sector
        result = map_sector("Energy", from_system="gics")
        assert result == "Energy"

    def test_gics_l2_oil_gas(self):
        from src.sector_mapper import map_sector
        result = map_sector("Oil, Gas & Consumable Fuels", from_system="gics_l2")
        assert result == "Energy"

    def test_nace_to_ngfs(self):
        from src.sector_mapper import map_sector
        result = map_sector("D35", from_system="nace")
        assert result == "Energy"

    def test_bmv_to_ngfs(self):
        from src.sector_mapper import map_sector
        result = map_sector("Energía", from_system="bmv")
        assert result == "Energy"

    def test_iea_to_ngfs(self):
        from src.sector_mapper import map_sector
        result = map_sector("Transport", from_system="iea")
        assert result == "Transport"

    def test_unknown_sector_returns_none(self):
        from src.sector_mapper import map_sector
        result = map_sector("ZZ Unknown Sector XYZ", from_system="gics")
        assert result is None

    def test_invalid_system_raises(self):
        from src.sector_mapper import map_sector
        with pytest.raises(ValueError):
            map_sector("Energy", from_system="bloomberg")

    def test_map_portfolio_sectors_adds_column(self):
        from src.sector_mapper import map_portfolio_sectors
        df = pd.DataFrame({
            "company_name": ["PEMEX", "Cemex"],
            "sector": ["Oil, Gas & Consumable Fuels", "Construction Materials"],
        })
        result = map_portfolio_sectors(df, "sector", from_system="gics_l2")
        assert "ngfs_sector" in result.columns
        assert result["ngfs_sector"].iloc[0] == "Energy"
        assert result["ngfs_sector"].iloc[1] == "Industry"

    def test_carbon_intensity_energy_highest(self):
        from src.sector_mapper import get_carbon_intensity
        energy_intensity = get_carbon_intensity("Energy")
        buildings_intensity = get_carbon_intensity("Buildings")
        assert energy_intensity > buildings_intensity

    def test_stranded_asset_nz2050_higher_than_cp(self):
        from src.sector_mapper import get_stranded_asset_risk
        nz = get_stranded_asset_risk("Energy", "Net Zero 2050")
        cp = get_stranded_asset_risk("Energy", "Current Policies")
        assert nz > cp


# ── TransitionScorer ──────────────────────────────────────────────────────────

class TestTransitionScorer:

    @pytest.fixture
    def scorer(self):
        from src.ngfs_loader import NGFSLoader
        from src.transition_scorer import TransitionScorer
        loader = NGFSLoader(use_api=False, use_cache=False)
        data = loader.load()
        return TransitionScorer(ngfs_data=data)

    def test_score_sector_returns_sector_trei(self, scorer):
        from src.transition_scorer import SectorTREI
        result = scorer.score_sector("Energy", "Net Zero 2050", "medium")
        assert isinstance(result, SectorTREI)

    def test_trei_in_valid_range(self, scorer):
        result = scorer.score_sector("Energy", "Net Zero 2050", "medium")
        assert 0 <= result.trei <= 100

    def test_all_component_scores_in_range(self, scorer):
        result = scorer.score_sector("Energy", "Net Zero 2050", "medium")
        assert 0 <= result.policy_risk_score <= 100
        assert 0 <= result.technology_risk_score <= 100
        assert 0 <= result.market_risk_score <= 100

    def test_energy_higher_than_financials_nz2050(self, scorer):
        energy = scorer.score_sector("Energy", "Net Zero 2050", "medium")
        fin = scorer.score_sector("Financials", "Net Zero 2050", "medium")
        assert energy.trei > fin.trei

    def test_nz2050_higher_trei_than_cp_for_energy(self, scorer):
        """Under NZ2050, transition risk for Energy should exceed Current Policies."""
        nz = scorer.score_sector("Energy", "Net Zero 2050", "medium")
        cp = scorer.score_sector("Energy", "Current Policies", "medium")
        assert nz.trei > cp.trei

    def test_risk_tier_assignment(self, scorer):
        result = scorer.score_sector("Energy", "Net Zero 2050", "medium")
        assert result.risk_tier in [
            "Very Low", "Low", "Medium", "High", "Very High"
        ]

    def test_score_all_sectors_shape(self, scorer):
        df = scorer.score_all_sectors()
        assert "Sector" in df.columns
        assert "TREI" in df.columns
        assert len(df) > 0

    def test_heatmap_data_matrix_shape(self, scorer):
        hm = scorer.heatmap_data()
        # Rows = sectors, Cols = scenarios
        assert hm.shape[0] >= 5
        assert hm.shape[1] >= 4

    def test_portfolio_trei_returns_portfolio_trei(self, scorer):
        from src.transition_scorer import PortfolioTREI
        weights = {"Energy": 0.4, "Transport": 0.3, "Industry": 0.3}
        result = scorer.score_portfolio(weights, "Net Zero 2050", "medium")
        assert isinstance(result, PortfolioTREI)
        assert 0 <= result.portfolio_trei <= 100

    def test_portfolio_trei_zero_weights_raises(self, scorer):
        with pytest.raises((ValueError, ZeroDivisionError)):
            scorer.score_portfolio({"Energy": 0.0}, "Net Zero 2050")

    def test_what_if_sbti_reduces_trei(self, scorer):
        weights = {"Energy": 0.5, "Industry": 0.3, "Transport": 0.2}
        baseline, adjusted = scorer.what_if_sbti(
            weights, "Net Zero 2050", sbti_adoption_pct=0.50
        )
        assert adjusted.portfolio_trei <= baseline.portfolio_trei

    def test_robustness_table_has_all_scenarios(self, scorer):
        weights = {"Energy": 0.4, "Buildings": 0.3, "Agriculture": 0.3}
        rob = scorer.robustness_table(weights)
        assert "Net Zero 2050" in rob.index
        assert "Current Policies" in rob.index


# ── PortfolioAnalyzer ─────────────────────────────────────────────────────────

class TestPortfolioAnalyzer:

    @pytest.fixture
    def analyzer(self):
        from src.ngfs_loader import NGFSLoader
        from src.portfolio_analyzer import PortfolioAnalyzer
        loader = NGFSLoader(use_api=False, use_cache=False)
        data = loader.load()
        return PortfolioAnalyzer(ngfs_data=data)

    def test_load_portfolio_returns_dataframe(self, analyzer):
        df = analyzer.load_portfolio()
        assert isinstance(df, pd.DataFrame)
        assert "company_name" in df.columns
        assert "ngfs_sector" in df.columns

    def test_sample_portfolio_has_40_companies(self, analyzer):
        df = analyzer.load_portfolio()
        assert len(df) == 40

    def test_weights_sum_to_one(self, analyzer):
        df = analyzer.load_portfolio()
        assert abs(df["weight"].sum() - 1.0) < 0.01

    def test_analyze_returns_result(self, analyzer):
        from src.portfolio_analyzer import PortfolioAnalysisResult
        result = analyzer.analyze()
        assert isinstance(result, PortfolioAnalysisResult)

    def test_coverage_pct_positive(self, analyzer):
        result = analyzer.analyze()
        assert result.coverage_pct > 0

    def test_sector_weights_sum_to_approx_one(self, analyzer):
        result = analyzer.analyze()
        total = sum(result.sector_weights.values())
        assert abs(total - 1.0) < 0.05

    def test_portfolio_trei_all_scenarios_present(self, analyzer):
        result = analyzer.analyze()
        assert "Net Zero 2050" in result.portfolio_trei
        assert "Current Policies" in result.portfolio_trei

    def test_sbti_impact_reduces_or_neutral(self, analyzer):
        result = analyzer.analyze()
        for scenario, impact in result.sbti_impact.items():
            assert impact >= 0, f"SBTi adoption should not increase TREI ({scenario})"

    def test_companies_list_populated(self, analyzer):
        result = analyzer.analyze()
        assert len(result.companies) > 0

    def test_companies_dataframe_has_expected_columns(self, analyzer):
        result = analyzer.analyze()
        df = analyzer.companies_dataframe(result)
        assert "TREI NZ2050" in df.columns
        assert "SBTi Status" in df.columns
        assert "Risk Flags" in df.columns


# ── GHGReporter ───────────────────────────────────────────────────────────────

class TestGHGReporter:

    @pytest.fixture
    def reporter_and_result(self, tmp_path):
        from src.ngfs_loader import NGFSLoader
        from src.portfolio_analyzer import PortfolioAnalyzer
        from src.reporter import GHGReporter
        loader = NGFSLoader(use_api=False, use_cache=False)
        data = loader.load()
        analyzer = PortfolioAnalyzer(ngfs_data=data)
        result = analyzer.analyze()
        reporter = GHGReporter(result=result, org_name="Test Org")
        return reporter, result, tmp_path

    def test_excel_report_created(self, reporter_and_result):
        reporter, result, tmp_path = reporter_and_result
        out = tmp_path / "test_report.xlsx"
        path = reporter.to_excel(out)
        assert path.exists()
        assert path.stat().st_size > 1000

    def test_html_report_created(self, reporter_and_result):
        reporter, result, tmp_path = reporter_and_result
        out = tmp_path / "test_report.html"
        path = reporter.to_html(out)
        assert path.exists()
        assert "IFRS S2" in path.read_text(encoding="utf-8")

    def test_html_contains_org_name(self, reporter_and_result):
        reporter, result, tmp_path = reporter_and_result
        out = tmp_path / "test_report.html"
        reporter.to_html(out)
        content = out.read_text(encoding="utf-8")
        assert "Test Org" in content

    def test_excel_has_expected_sheets(self, reporter_and_result):
        reporter, result, tmp_path = reporter_and_result
        import openpyxl
        out = tmp_path / "test_report.xlsx"
        reporter.to_excel(out)
        wb = openpyxl.load_workbook(out)
        assert "Cover" in wb.sheetnames
        assert "Methodology" in wb.sheetnames


# ── Integration test ──────────────────────────────────────────────────────────

class TestEndToEnd:

    def test_full_pipeline_runs(self, tmp_path):
        """Full pipeline from data load to report export."""
        from src.ngfs_loader import NGFSLoader
        from src.transition_scorer import TransitionScorer
        from src.portfolio_analyzer import PortfolioAnalyzer
        from src.reporter import GHGReporter

        loader = NGFSLoader(use_api=False, use_cache=False)
        ngfs_data = loader.load()

        scorer = TransitionScorer(ngfs_data=ngfs_data)
        analyzer = PortfolioAnalyzer(ngfs_data=ngfs_data)

        result = analyzer.analyze(horizons=["medium"])

        reporter = GHGReporter(result=result, org_name="E2E Test", scorer=scorer)
        excel_path = reporter.to_excel(tmp_path / "e2e.xlsx")
        html_path = reporter.to_html(tmp_path / "e2e.html")

        assert excel_path.exists()
        assert html_path.exists()

    def test_sector_heatmap_values_bounded(self):
        """All TREI values in heatmap should be in [0, 100]."""
        from src.ngfs_loader import NGFSLoader
        from src.transition_scorer import TransitionScorer

        loader = NGFSLoader(use_api=False, use_cache=False)
        data = loader.load()
        scorer = TransitionScorer(ngfs_data=data)

        hm = scorer.heatmap_data("medium")
        assert hm.values.min() >= 0
        assert hm.values.max() <= 100
